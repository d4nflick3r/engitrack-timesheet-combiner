import streamlit as st
import io
import os
import sys
import base64
import datetime
import subprocess
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── PWA patch (keeps the hosted web-app manifest working) ────────────────────
try:
    import patch_pwa
    patch_pwa.patch()
except Exception:
    pass

st.set_page_config(
    page_title="EngiTrack Timesheet Combiner",
    page_icon="📋",
    layout="wide",
)

st.title("EngiTrack Timesheet Combiner")
st.markdown(
    "Upload **SOSengitrack** weekly CSV exports for one or more engineers. "
    "The app combines them into a single Excel workbook with **Weekly** and **Monthly** totals."
)
st.divider()


# ── CSV Parser ───────────────────────────────────────────────────────────────

def _parse_date(s):
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s.strip(), fmt).date()
        except Exception:
            pass
    return None


_WEEKDAYS = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday"}


def parse_csv(file_obj):
    """Parse a SOSengitrack weekly timesheet CSV and return a data dict.

    Weekday OT is recalculated from daily rows as max(0, hours - 9) per
    weekday (Mon–Fri, excluding Bank Holiday days), overriding the CSV total.
    """
    raw = file_obj.read()
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8-sig")
    lines = raw.splitlines()

    data = {
        "engineer": "",
        "week_str": "",
        "week_date": None,
        "total_hours": 0.0,
        "standard_hours": 0.0,
        "weekday_ot": 0.0,
        "saturday_hours": 0.0,
        "sunday_hours": 0.0,
        "bh_hours": 0.0,
        "sick_days": 0,
        "repairs": 0,
        "extra_jobs": 0,
    }

    # Collect per-day records for OT recalculation
    daily_rows = []   # list of (day_name, hours, is_bh, is_sick)

    in_daily = False
    in_totals = False

    for line in lines:
        s = line.strip()

        # Header fields
        if s.startswith("Engineer,"):
            data["engineer"] = s.split(",", 1)[1].strip()
            continue
        if s.startswith("Week Commencing,"):
            wc = s.split(",", 1)[1].strip()
            data["week_str"] = wc
            data["week_date"] = _parse_date(wc)
            continue

        # Section markers
        if s.startswith("Date,Day,"):
            in_daily = True
            in_totals = False
            continue
        if s == "Weekly Totals":
            in_daily = False
            in_totals = True
            continue
        if not s:
            in_daily = False
            continue

        # Daily data rows
        # Columns: Date,Day,Start,End,Total Hours,Bank Holiday,Holiday,Sickness,Weekend Worked
        if in_daily:
            parts = s.split(",")
            if len(parts) >= 8:
                day_name  = parts[1].strip()
                try:
                    hours = float(parts[4].strip()) if parts[4].strip() else 0.0
                except ValueError:
                    hours = 0.0
                is_bh   = parts[5].strip().lower() == "yes"
                is_sick = parts[7].strip().lower() == "yes"
                daily_rows.append((day_name, hours, is_bh, is_sick))
                if is_sick:
                    data["sick_days"] += 1
            continue

        # Weekly totals / repairs / extra jobs
        if "," in s:
            key, val = s.split(",", 1)
            key = key.strip()
            val = val.strip().strip('"')
            try:
                v = float(val)
            except ValueError:
                v = 0.0
            if key == "Total Hours":
                data["total_hours"] = v
            elif "Standard Hours" in key:
                data["standard_hours"] = v
            elif key == "Saturday Hours":
                data["saturday_hours"] = v
            elif key == "Sunday Hours":
                data["sunday_hours"] = v
            elif key == "Bank Holiday Hours":
                data["bh_hours"] = v
            elif key == "Repairs Logged":
                data["repairs"] = int(v)
            elif key == "Extra Jobs Logged":
                data["extra_jobs"] = int(v)
            # "Weekday Overtime" from CSV is intentionally ignored —
            # we recalculate it below.

    # ── Recalculate weekday OT ────────────────────────────────────────────────
    # Rules (both must be satisfied):
    #   1. A day must exceed 9 h before any OT is counted for that day.
    #   2. Salary covers 40 h/week — OT is only recorded once 40 h are reached.
    #
    # Method: standard hours = min(daily, 9) per weekday (non-BH).
    # Any salary bandwidth unused by standard hours (40 − standard_hours)
    # absorbs the first N hours of per-day OT before OT is recorded.
    weekday_rows = [
        (hours, is_bh)
        for day_name, hours, is_bh, is_sick in daily_rows
        if day_name in _WEEKDAYS
    ]
    standard_hours = sum(min(h, 9.0) for h, is_bh in weekday_rows if not is_bh)
    per_day_ot     = sum(max(0.0, h - 9.0) for h, is_bh in weekday_rows if not is_bh)
    salary_deficit = max(0.0, 40.0 - standard_hours)   # unused salary bandwidth
    data["weekday_ot"] = round(max(0.0, per_day_ot - salary_deficit), 2)

    return data


# ── Excel builder ─────────────────────────────────────────────────────────────

_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _hdr(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = _BORDER
    return c


def _cell(ws, row, col, value, bold=False, center=True, num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, size=11)
    c.alignment = Alignment(horizontal="center" if center else "left", vertical="center")
    c.border = _BORDER
    if num_fmt:
        c.number_format = num_fmt
    return c


def _total_cell(ws, row, col, value, center=True):
    c = _cell(ws, row, col, value, bold=True, center=center)
    c.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    return c


WEEKLY_HEADERS = [
    "Engineer", "Week Commencing",
    "Total Hours", "Standard Hours", "Weekday OT",
    "Sat Hours", "Sun Hours", "BH Hours",
    "Sick Days", "Repairs", "Extra Jobs",
]

MONTHLY_HEADERS = [
    "Engineer", "Month",
    "Total Hours", "Standard Hours", "Weekday OT",
    "Sat Hours", "Sun Hours", "BH Hours",
    "Sick Days", "Repairs", "Extra Jobs",
]

NUMERIC_COLS = [
    "total_hours", "standard_hours", "weekday_ot",
    "saturday_hours", "sunday_hours", "bh_hours",
    "sick_days", "repairs", "extra_jobs",
]


def _weekly_row(ts):
    return [
        ts["engineer"],
        ts["week_date"] if ts["week_date"] else ts["week_str"],
        ts["total_hours"] or None,
        ts["standard_hours"] or None,
        ts["weekday_ot"] or None,
        ts["saturday_hours"] or None,
        ts["sunday_hours"] or None,
        ts["bh_hours"] or None,
        ts["sick_days"] or None,
        ts["repairs"] or None,
        ts["extra_jobs"] or None,
    ]


def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel(timesheets):
    timesheets = sorted(
        timesheets,
        key=lambda t: (t["engineer"].lower(), t["week_date"] or datetime.date.min),
    )

    wb = Workbook()
    wb.remove(wb.active)

    # ── Weekly sheet ──────────────────────────────────────────────────────────
    ws_w = wb.create_sheet("Weekly")
    ws_w.sheet_view.showGridLines = False
    ws_w.row_dimensions[1].height = 28

    for col, h in enumerate(WEEKLY_HEADERS, 1):
        _hdr(ws_w, 1, col, h)

    for r, ts in enumerate(timesheets, 2):
        row = _weekly_row(ts)
        for col, val in enumerate(row, 1):
            fmt = "DD/MM/YYYY" if col == 2 and isinstance(val, datetime.date) else None
            _cell(ws_w, r, col, val, center=(col > 1), num_fmt=fmt)
        ws_w.row_dimensions[r].height = 18

    # Total row
    tr = len(timesheets) + 2
    _total_cell(ws_w, tr, 1, "TOTAL", center=False)
    for col in range(2, len(WEEKLY_HEADERS) + 1):
        vals = [ws_w.cell(row=r, column=col).value for r in range(2, tr)]
        nums = [v for v in vals if isinstance(v, (int, float))]
        _total_cell(ws_w, tr, col, round(sum(nums), 2) if nums else None)
    ws_w.row_dimensions[tr].height = 20

    _set_widths(ws_w, [22, 16, 13, 14, 12, 11, 11, 11, 11, 11, 11])
    ws_w.freeze_panes = "A2"

    # ── Monthly sheet ─────────────────────────────────────────────────────────
    ws_m = wb.create_sheet("Monthly")
    ws_m.sheet_view.showGridLines = False
    ws_m.row_dimensions[1].height = 28

    for col, h in enumerate(MONTHLY_HEADERS, 1):
        _hdr(ws_m, 1, col, h)

    # Aggregate by (engineer, year, month)
    monthly = {}
    for ts in timesheets:
        wd = ts["week_date"]
        if wd:
            key = (ts["engineer"], wd.year, wd.month)
            label = wd.strftime("%B %Y")
        else:
            key = (ts["engineer"], 0, 0)
            label = "Unknown"

        if key not in monthly:
            monthly[key] = {
                "engineer": ts["engineer"],
                "label": label,
                "total_hours": 0.0,
                "standard_hours": 0.0,
                "weekday_ot": 0.0,
                "saturday_hours": 0.0,
                "sunday_hours": 0.0,
                "bh_hours": 0.0,
                "sick_days": 0,
                "repairs": 0,
                "extra_jobs": 0,
            }
        d = monthly[key]
        for f in NUMERIC_COLS:
            d[f] += ts[f]

    sorted_keys = sorted(monthly, key=lambda k: (k[0].lower(), k[1], k[2]))

    for r, key in enumerate(sorted_keys, 2):
        d = monthly[key]
        row = [
            d["engineer"], d["label"],
            d["total_hours"] or None,
            d["standard_hours"] or None,
            d["weekday_ot"] or None,
            d["saturday_hours"] or None,
            d["sunday_hours"] or None,
            d["bh_hours"] or None,
            d["sick_days"] or None,
            d["repairs"] or None,
            d["extra_jobs"] or None,
        ]
        for col, val in enumerate(row, 1):
            _cell(ws_m, r, col, val, center=(col > 1))
        ws_m.row_dimensions[r].height = 18

    # Total row
    tr_m = len(sorted_keys) + 2
    _total_cell(ws_m, tr_m, 1, "TOTAL", center=False)
    for col in range(2, len(MONTHLY_HEADERS) + 1):
        vals = [ws_m.cell(row=r, column=col).value for r in range(2, tr_m)]
        nums = [v for v in vals if isinstance(v, (int, float))]
        _total_cell(ws_m, tr_m, col, round(sum(nums), 2) if nums else None)
    ws_m.row_dimensions[tr_m].height = 20

    _set_widths(ws_m, [22, 16, 13, 14, 12, 11, 11, 11, 11, 11, 11])
    ws_m.freeze_panes = "A2"

    return wb


# ── UI ────────────────────────────────────────────────────────────────────────

uploaded_files = st.file_uploader(
    "Upload SOSengitrack weekly timesheet CSV files",
    type=["csv"],
    accept_multiple_files=True,
    help="Upload one or more SOSengitrack weekly CSV exports — any number of engineers or weeks.",
)

if uploaded_files:
    timesheets = []
    errors = []

    for f in uploaded_files:
        try:
            f.seek(0)
            ts = parse_csv(f)
            if not ts["engineer"]:
                errors.append(f"{f.name}: Could not find engineer name — is this a SOSengitrack CSV?")
            else:
                timesheets.append(ts)
        except Exception as exc:
            errors.append(f"{f.name}: {exc}")

    for err in errors:
        st.error(err)

    if timesheets:
        engineers = sorted(set(ts["engineer"] for ts in timesheets))
        weeks = sorted(set(ts["week_str"] for ts in timesheets))

        col1, col2, col3 = st.columns(3)
        col1.metric("Files loaded", len(timesheets))
        col2.metric("Engineers", len(engineers))
        col3.metric("Weeks", len(weeks))

        with st.expander("Preview loaded timesheets"):
            for ts in sorted(timesheets, key=lambda t: (t["engineer"].lower(), t["week_str"])):
                st.markdown(
                    f"**{ts['engineer']}** · week of {ts['week_str']} · "
                    f"{ts['total_hours']}h total, {ts['weekday_ot']}h OT, "
                    f"{ts['bh_hours']}h BH, {ts['sick_days']} sick day(s)"
                )

        st.divider()

        if st.button("Combine into Excel Workbook", type="primary"):
            wb = build_excel(timesheets)
            buf = io.BytesIO()
            wb.save(buf)
            st.session_state["excel_bytes"] = buf.getvalue()

        if st.session_state.get("excel_bytes"):
            excel_bytes = st.session_state["excel_bytes"]
            fname = "timesheets_combined.xlsx"

            # Windows exe — save directly to Downloads folder
            if getattr(sys, "frozen", False):
                ts_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                unique_fname = f"timesheets_combined_{ts_str}.xlsx"
                downloads = Path.home() / "Downloads" / unique_fname
                try:
                    downloads.write_bytes(excel_bytes)
                    st.success(f"Saved to {downloads}")
                    if st.button("Open file location"):
                        subprocess.Popen(f'explorer /select,"{downloads}"')
                except PermissionError:
                    st.error(
                        f"Could not save to Downloads — close any open copy of the file in Excel and try again.\n\n"
                        f"Attempted path: {downloads}"
                    )
            else:
                b64 = base64.b64encode(excel_bytes).decode()
                st.markdown(
                    f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" '
                    f'download="{fname}" style="font-size:1.1em;font-weight:600;">⬇ Download Combined Workbook</a>',
                    unsafe_allow_html=True,
                )
                st.success(
                    f"Workbook ready — {len(timesheets)} timesheet(s), 2 sheets: **Weekly** · **Monthly**"
                )
                st.caption(
                    "If the download link doesn't work, open this page in a new browser tab "
                    "(not the embedded preview)."
                )

else:
    st.info("Upload SOSengitrack weekly timesheet CSV files above to get started.")
    st.markdown("""
**How to use:**
1. Export each engineer's weekly timesheet from SOSengitrack as a CSV
2. Upload all files at once (any number of engineers, any number of weeks)
3. Click **Combine into Excel Workbook**
4. Download the file — it contains:
   - **Weekly** sheet: one row per engineer per week with all hour totals
   - **Monthly** sheet: the same data rolled up by calendar month
""")
